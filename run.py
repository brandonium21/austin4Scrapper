import openpyxl
import requests
from bs4 import BeautifulSoup
import csv
import sys

f = open('austinResults.csv', 'a')
try:
    writer = csv.writer(f)
    writer.writerow( ('Name', 'Website Address', 'Verdict') )
finally:
    f.close()

wb = openpyxl.load_workbook('spreadsheet.xlsx')

ws = wb.get_sheet_by_name('IA_SEC_-_FIRM_ROSTER_FOIA_DOWNL')
weblist = []
namelist = []
final = []
for row in ws.iter_rows('AA{}:AA{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        weblist.append(cell.value)
print 'Done Collecting Websites'


for row in ws.iter_rows('E{}:E{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        namelist.append(cell.value)
print 'Done Collecting Names'

def check(link_text):
    verdict = []
    checklist = ["guaranteed return", "no losses", "guaranteed principal", "cannot lose money", "no risk", "never lose money", "guaranteed to not lose money", "riskless return", "easy and safe return"]
    for item in checklist:
        if item in link_text.lower():
            verdict.append(item)
    e_v = ",".join(verdict)
    if not verdict:
        e_v = 'None'
    return e_v

for index, item in enumerate(namelist):
    if index == 0:
        continue
    print 'Evaluating {}'.format(item)
    if not weblist[index] or not weblist[index].lower().startswith('http') or weblist[index].lower().startswith('mailto'):
        continue
    soup = BeautifulSoup(requests.get(weblist[index], verify=False).text, 'html.parser')
    site_text = soup.get_text()
    for link in soup.find_all('a'):
        if link.get('href'):
            print link
            if (link.get('href').lower().startswith('http')):
                loup = BeautifulSoup(requests.get(link.get('href'), verify=False).text, 'html.parser').get_text()
                site_text += ' '
                site_text += loup
    result = check(site_text)

    f = open('austinResults.csv', 'a')
    try:
        writer = csv.writer(f)
        writer.writerow( (item, weblist[index], result) )
    finally:
        f.close()
